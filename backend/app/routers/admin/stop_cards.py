import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.stop_card import StopCard, StopCardStatus
from app.models.user import User
from app.repositories.stop_card import StopCardRepository
from app.repositories.stop_card_photo import StopCardPhotoRepository
from app.repositories.user import UserRepository
from app.routers.deps import require_admin
from app.schemas.stop_card import StopCardResponse
from app.services.stop_card import StopCardService

router = APIRouter(prefix="/stop-cards", tags=["admin-stop-cards"])


def _service(db: AsyncSession = Depends(get_db)) -> StopCardService:
    return StopCardService(
        StopCardRepository(db),
        StopCardPhotoRepository(db),
        UserRepository(db),
    )


@router.get("", response_model=list[StopCardResponse])
async def list_stop_cards(
    section_id: int | None = Query(None),
    card_status: StopCardStatus | None = Query(None, alias="status"),
    year: int | None = Query(None),
    month: int | None = Query(None),
    svc: StopCardService = Depends(_service),
    _: User = Depends(require_admin),
):
    if year and month:
        cards = await svc.get_by_month(year, month)
    elif section_id:
        cards = await svc.get_by_section(section_id)
    else:
        cards = await svc.repo.get_all()

    if year and month and section_id:
        cards = [c for c in cards if c.section_id == section_id]
    if card_status:
        cards = [c for c in cards if c.status == card_status]
    return cards


STATUS_LABELS = {
    "created": "Создана",
    "waiting_violator": "Ожидает регистрации нарушителя",
    "violator_fixing": "Устраняется нарушителем",
    "manager_review": "Проверка менеджера",
    "safety_check": "Проверка ОТ и ТБ",
    "approved": "Разрешено к работе",
    "rejected": "Запрещено",
    "closed": "Закрыто",
}

HEADERS = [
    "№", "Нарушитель", "Участок (ID)", "Описание", "Статус", "Наблюдатель",
    "Дата создания",
    "Принял (кто)", "Принял (когда)",
    "Описание устранения", "Устранил (кто)", "Устранил (когда)",
    "Заметка менеджера", "Проверил менеджер (кто)", "Проверил менеджер (когда)",
    "Заметка ОТ и ТБ", "Проверил ОТ и ТБ (кто)", "Проверил ОТ и ТБ (когда)",
    "Закрыта",
]


def _fmt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y %H:%M")


def _build_excel(cards: list[StopCard]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Стоп-карты"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 36

    for row_idx, c in enumerate(cards, start=2):
        row = [
            c.id,
            c.violator_name,
            c.section_id,
            c.description,
            STATUS_LABELS.get(c.status.value, c.status.value),
            c.reporter.full_name if c.reporter else "",
            _fmt(c.created_at),
            c.acknowledged_by.full_name if c.acknowledged_by else "",
            _fmt(c.acknowledged_at),
            c.fix_description or "",
            c.fixed_by.full_name if c.fixed_by else "",
            _fmt(c.fixed_at),
            c.manager_note or "",
            c.manager_checked_by.full_name if c.manager_checked_by else "",
            _fmt(c.manager_checked_at),
            c.safety_note or "",
            c.safety_checked_by.full_name if c.safety_checked_by else "",
            _fmt(c.safety_checked_at),
            _fmt(c.closed_at),
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [6, 24, 12, 40, 28, 24, 18, 24, 18, 40, 24, 18, 30, 24, 18, 30, 24, 18, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/export")
async def export_stop_cards(
    section_id: int | None = Query(None),
    card_status: StopCardStatus | None = Query(None, alias="status"),
    year: int | None = Query(None),
    month: int | None = Query(None),
    svc: StopCardService = Depends(_service),
    _: User = Depends(require_admin),
):
    if year and month:
        cards = await svc.get_by_month(year, month)
    elif section_id:
        cards = await svc.get_by_section(section_id)
    else:
        cards = await svc.repo.get_all()

    if year and month and section_id:
        cards = [c for c in cards if c.section_id == section_id]
    if card_status:
        cards = [c for c in cards if c.status == card_status]

    content = _build_excel(cards)
    filename = f"stop_cards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{stop_card_id}", response_model=StopCardResponse)
async def get_stop_card(
    stop_card_id: int,
    svc: StopCardService = Depends(_service),
    _: User = Depends(require_admin),
):
    try:
        return await svc.get_by_id(stop_card_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.patch("/{stop_card_id}/close", response_model=StopCardResponse)
async def close_stop_card(
    stop_card_id: int,
    svc: StopCardService = Depends(_service),
    _: User = Depends(require_admin),
):
    try:
        return await svc.close(stop_card_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
